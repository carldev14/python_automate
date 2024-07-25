from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import calendar

def main():
    # Create a new workbook
    wb = Workbook()

    # Get the active worksheet
    ws = wb.active

    months = [calendar.month_name[i] for i in range(1, 13)]  # Get names of months

    for month in months:
        ws = wb.create_sheet(title=month)

        # Append data to the worksheet
        ws.append(['Date', 'Company Name', 'Address', 'TIN', 'Description',
                   'Total Amount Paid', 'Input Vat', 'Cost of Products & Services', 'Remarks'])
        ws.append(['', '', '', '', '', '', '', '', ''])

        # Merge cells
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws.merge_cells('C1:C2')
        ws.merge_cells('D1:D2')
        ws.merge_cells('E1:E2')
        ws.merge_cells('F1:F2')
        ws.merge_cells('G1:G2')
        ws.merge_cells('H1:H2')
        ws.merge_cells('I1:I2')

        # Set column dimensions
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20

        # Create a fill object with a background color
        fill = PatternFill(start_color='666666', end_color='666666', fill_type='solid')

        # Create a font object with a text color
        font = Font(color='FFFFFF', bold=True, name="Arial", size=10)  # White text color

        # Create an alignment object to center the text
        alignment = Alignment(horizontal='center', vertical='center')

        # Set the fill, font, and alignment properties for each cell in the first two rows
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=9):
            for cell in row:
                cell.fill = fill
                cell.font = font
                cell.alignment = alignment

    SaveAs = input('Enter the file name including the path where you want to save it (e.g., "/home/carl/python/sample.xlsx"): ')
    wb.save(SaveAs)

if __name__ == "__main__":
    main()
